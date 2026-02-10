from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# helpers
# -----------------------------
def _norm_text(s: str) -> str:
    s = (s or "").strip().casefold().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s

def _to_address(v: Any) -> str:
    """
    Менеджерский ключ адреса:
    - оставляем: Город + Улица/локация + Базовый дом
    - убираем: пом/офис/кв/подъезд/этаж/пункты, корп/к/строения, перечисления 1/1,1/2, 39/3,39/4
    - убираем лишний "Тюмень," если дальше указан другой населённый пункт
    """
    s = str(v or "").strip().casefold().replace("ё", "е")
    if not s:
        return ""

    # нормализуем разделители (оставляем запятые)
    s = s.replace("—", "-").replace("–", "-")
    s = re.sub(r"[;:]+", ",", s)
    s = re.sub(r"\s*,\s*", ",", s)
    s = re.sub(r"\s+", " ", s)
    s = s.strip(" ,")

    parts = [p.strip() for p in s.split(",") if p.strip()]
    if not parts:
        return ""

    # если "тюмень, <другой город>, ..." — выкидываем "тюмень"
    if len(parts) >= 2:
        first = parts[0]
        second = parts[1]
        # второй кусок "похож на город": нет цифр
        if first == "тюмень" and not any(ch.isdigit() for ch in second):
            parts = parts[1:]

    # словари/замены
    def norm_piece(p: str) -> str:
        p = p.replace(".", "")
        p = re.sub(r"\b(г|город)\b", "", p).strip()
        # унификация
        rep = {
            "улица": "ул",
            "ул": "ул",
            "проспект": "пр-кт",
            "пр-т": "пр-кт",
            "переулок": "пер",
            "шоссе": "ш",
            "бульвар": "бул",
            "площадь": "пл",
            "микрорайон": "мкр",
            "мкрн": "мкр",
            "мкр": "мкр",
            "строение": "стр",
            "стр": "стр",
            "здание": "зд",
            "зд": "зд",
        }
        # заменяем слова по границам
        for a, b in rep.items():
            p = re.sub(rf"\b{re.escape(a)}\b", b, p)
        p = re.sub(r"\s+", " ", p).strip()
        return p

    # фильтры: выкидываем помещения/офисы/квартиры/подъезды/этажи/пункты
    def is_junk_piece(p: str) -> bool:
        return bool(re.search(r"\b(пом|помещение|офис|оф|кв|квартира|подъезд|п-?зд|эт|этаж|пункт|п)\b", p))

    parts = [norm_piece(p) for p in parts]
    parts = [p for p in parts if p and not is_junk_piece(p)]

    if not parts:
        return ""

    city = parts[0]

    # остаток считаем "улица/локация + дом"
    rest = parts[1:] if len(parts) > 1 else []

    # соберём текст остатка одной строкой
    rest_text = " ".join(rest)
    rest_text = re.sub(r"\s*/\s*", "/", rest_text)

    # выкидываем корп/к полностью (они часто дают ложные расхождения)
    rest_text = re.sub(r"\b(корпус|корп|к)\s*[\w/.-]+", "", rest_text)
    # выкидываем стр/зд (оставим только номер дома)
    rest_text = re.sub(r"\b(стр|зд)\s*[\w/.-]+", "", rest_text)

    rest_text = re.sub(r"\s+", " ", rest_text).strip()

    # дом: ищем первый "нормальный" номер дома (цифры + буква), и режем / и - хвосты:
    # 39/3 -> 39 ; 16а-1 -> 16а ; 576/7 -> 576
    house = ""
    m = re.search(r"\b(\d+[а-яa-z]?)((?:/\d+)|(?:-\d+))?\b", rest_text)
    if m:
        house = m.group(1)  # базовый дом без / или - части

    # улица/локация: удалим найденный дом (с хвостом, если был) и почистим
    street = rest_text
    if m:
        street = (rest_text[: m.start()] + " " + rest_text[m.end():]).strip()

    # ещё раз чистим мусорные хвосты перечислений
    street = re.sub(r"\b\d+(?:\s*,\s*\d+)+\b", "", street)
    street = re.sub(r"\s+", " ", street).strip(" ,")

    # финальная сборка
    if house:
        if street:
            return f"{city}, {street}, {house}".strip(" ,")
        return f"{city}, {house}".strip(" ,")

    # если дом не нашли — возвращаем хотя бы город + остаток
    if street:
        return f"{city}, {street}".strip(" ,")
    return city


def _to_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None

    if isinstance(v, date) and not isinstance(v, datetime):
        return v

    if isinstance(v, datetime):
        return v.date()

    if isinstance(v, str):
        s = v.strip()
        for fmt in (
            "%d.%m.%Y",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d.%m.%y",
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M",
        ):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue

    return None


def _to_hours(v: Any) -> Optional[Decimal]:
    if v is None or v == "":
        return None

    if isinstance(v, str) and v.strip() in ("-", "—"):
        return None

    # время (02:30) -> часы
    if isinstance(v, time):
        return (Decimal(v.hour) + (Decimal(v.minute) / Decimal(60))).quantize(Decimal("0.01"))

    if isinstance(v, datetime):
        t = v.time()
        return (Decimal(t.hour) + (Decimal(t.minute) / Decimal(60))).quantize(Decimal("0.01"))

    try:
        if isinstance(v, (int, float, Decimal)):
            return Decimal(str(v)).quantize(Decimal("0.01"))
        if isinstance(v, str):
            s = v.strip().replace(",", ".")
            return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None

    return None


def _header_to_date(raw: Any, default_year: int) -> Optional[date]:
    """
    Даты в шапке у заказчика бывают:
    - как date/datetime (Excel дата)
    - как строка "1.1"
    """
    d = _to_date(raw)
    if d:
        return d

    if isinstance(raw, str):
        s = raw.strip()
        m = re.fullmatch(r"(\d{1,2})\.(\d{1,2})", s)
        if m:
            dd = int(m.group(1))
            mm = int(m.group(2))
            return date(default_year, mm, dd)

    return None


def _find_header_map(header_row: Iterable[Any]) -> Dict[str, int]:
    """
    Заголовки для выгрузки с базы.
    """
    idx: Dict[str, int] = {}

    for i, raw in enumerate(header_row):
        h = _norm_text(str(raw or ""))
        if not h:
            continue

        # дата
        if h == "дата" or h.startswith("дата ") or h == "день" or h == "период":
            idx.setdefault("date", i)

        # часы
        if "час" in h or "отработ" in h or "hours" in h:
            idx.setdefault("hours", i)

        # адрес
        if h == "сегмент" or "адрес" in h or "address" in h:
            idx.setdefault("address", i)

        # город
        if h == "город":
            idx.setdefault("city", i)

    return idx


# -----------------------------
# Variant B: safe address merge (fuzzy)
# -----------------------------
_SUFFIX_RE = re.compile(r"(\s*/\s*\d+|\s*-\s*\d+|\s+к\s*\d+|\s+стр\s*\d+)\s*$")


def _base_addr(addr_norm: str) -> str:
    """
    База адреса для склейки:
    - убираем конечные суффиксы типа "/7", "-1", "к 2", "стр 1"
    Только 1 шаг (чтобы не переусердствовать).
    """
    a = (addr_norm or "").strip()
    return _SUFFIX_RE.sub("", a).strip()


def _build_safe_merge_map(customer_keys: set, db_keys: set) -> Dict[str, str]:
    """
    Возвращает mapping {variant_addr -> base_addr}, но только когда это безопасно:
    - по base_addr есть РОВНО 2 варианта
    - один из вариантов = base_addr
    - второй вариант = base_addr + суффикс (/, -, к, стр)
    - варианты разнесены по источникам (один только в customer, другой только в db)
    """
    all_addrs = {k.addr for k in customer_keys} | {k.addr for k in db_keys}

    base_to_variants: Dict[str, set] = {}
    for a in all_addrs:
        b = _base_addr(a)
        base_to_variants.setdefault(b, set()).add(a)

    mapping: Dict[str, str] = {}

    for b, vars_ in base_to_variants.items():
        if len(vars_) != 2:
            continue
        if b not in vars_:
            continue

        other = next(v for v in vars_ if v != b)

        # убеждаемся, что other реально "b + суффикс"
        if _base_addr(other) != b or not _SUFFIX_RE.search(other):
            continue

        b_in_c = any(k.addr == b for k in customer_keys)
        o_in_c = any(k.addr == other for k in customer_keys)
        b_in_d = any(k.addr == b for k in db_keys)
        o_in_d = any(k.addr == other for k in db_keys)

        # безопасный кейс:
        # customer: other, db: b (или наоборот)
        separated = (o_in_c and b_in_d and not b_in_c and not o_in_d) or (
            b_in_c and o_in_d and not o_in_c and not b_in_d
        )
        if not separated:
            continue

        # маппим "длинный" вариант -> base
        mapping[other] = b

    return mapping


def _apply_safe_address_merge(
    customer_rows: List["ParsedRow"], db_rows: List["ParsedRow"]
) -> Tuple[List["ParsedRow"], List["ParsedRow"]]:
    """
    Применяем безопасную склейку адресов (Variant B) к двум спискам строк.
    """
    customer_keys = {r.key for r in customer_rows}
    db_keys = {r.key for r in db_rows}

    addr_map = _build_safe_merge_map(customer_keys, db_keys)
    if not addr_map:
        return customer_rows, db_rows

    def remap(rows: List["ParsedRow"]) -> List["ParsedRow"]:
        out: List["ParsedRow"] = []
        for r in rows:
            new_addr = addr_map.get(r.key.addr, r.key.addr)
            if new_addr == r.key.addr:
                out.append(r)
                continue
            out.append(
                ParsedRow(
                    key=RowKey(d=r.key.d, addr=new_addr),
                    addr_raw=r.addr_raw,
                    hours=r.hours,
                    raw=r.raw,
                )
            )
        return out

    return remap(customer_rows), remap(db_rows)


# -----------------------------
# domain
# -----------------------------
@dataclass(frozen=True)
class RowKey:
    d: date
    addr: str  # нормализованный адрес


@dataclass
class ParsedRow:
    key: RowKey              # key.addr = НОРМАЛИЗОВАННЫЙ адрес (для сравнения)
    addr_raw: str            # адрес как в исходном файле
    hours: Decimal
    raw: Tuple[Any, ...]     # для отладки


# -----------------------------
# parsers
# -----------------------------
def parse_xlsx_matrix(path: str) -> List[ParsedRow]:
    """
    Файл заказчика (матрица):
    - строка: магазин (адрес)
    - колонки: дни (в шапке как Excel-даты или "1.1")
    - ячейки: часы
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]

    # колонка адреса
    addr_idx = None
    for i, raw in enumerate(header):
        if "адрес" in _norm_text(str(raw or "")):
            addr_idx = i
            break
    if addr_idx is None:
        raise ValueError("Не найдена колонка адреса в файле заказчика (матрица).")

    # колонки дат
    default_year = datetime.now().year
    day_cols: List[Tuple[int, date]] = []
    for i, raw in enumerate(header):
        d = _header_to_date(raw, default_year)
        if d:
            day_cols.append((i, d))

    if len(day_cols) < 10:
        raise ValueError("Не найдены колонки дат в файле заказчика (ожидается много дат в шапке).")

    out: List[ParsedRow] = []
    for r in rows[1:]:
        addr_raw = str(r[addr_idx] if addr_idx < len(r) else "").strip()
        addr_norm = _to_address(addr_raw)
        if not addr_norm:
            continue

        for col_i, d in day_cols:
            v = r[col_i] if col_i < len(r) else None
            h = _to_hours(v)
            if h is None or h == Decimal("0.00"):
                continue

            out.append(
                ParsedRow(
                    key=RowKey(d=d, addr=addr_norm),
                    addr_raw=addr_raw,
                    hours=h,
                    raw=tuple(r),
                )
            )

    return out


def _detect_header_row(rows: List[Tuple[Any, ...]], max_scan: int = 200) -> int:
    """
    В выгрузке с базы шапка может быть не первой строкой.
    Ищем строку, где есть: "дата" + ("сегмент"/"адрес") + "час"
    """
    scan = rows[: max_scan or 0]

    for i, r in enumerate(scan):
        texts = [_norm_text(str(x or "")) for x in r]
        joined = " ".join(texts)
        if ("дата" in joined) and (("сегмент" in joined) or ("адрес" in joined)) and ("час" in joined):
            return i

    # fallback: скоринг
    best_i = 0
    best_score = -1
    for i, r in enumerate(scan):
        texts = [_norm_text(str(x or "")) for x in r]
        score = 0
        if any(t == "дата" or t.startswith("дата ") for t in texts):
            score += 1
        if any("час" in t for t in texts):
            score += 1
        if any(t == "сегмент" or "адрес" in t for t in texts):
            score += 1
        if score > best_score:
            best_score = score
            best_i = i
        if score == 3:
            return i

    return best_i


def parse_xlsx_rowwise(path: str) -> List[ParsedRow]:
    """
    Выгрузка с базы (строки):
    - Дата
    - Город
    - Сегмент (адрес без города) / либо Адрес
    - Кол-во часов / Часы
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_i = _detect_header_row(rows, max_scan=250)
    header = rows[header_i]
    col = _find_header_map(header)

    # если дата не нашлась по заголовку — ищем по содержимому
    if "date" not in col:
        best_idx = None
        best_hits = 0
        sample = rows[header_i + 1 : header_i + 120]
        for j in range(len(header)):
            hits = 0
            for r in sample:
                v = r[j] if j < len(r) else None
                if _to_date(v) is not None:
                    hits += 1
            if hits > best_hits:
                best_hits = hits
                best_idx = j

        if best_idx is not None and best_hits >= 3:
            col["date"] = best_idx
        else:
            raise ValueError("Не удалось определить колонку даты в выгрузке.")

    missing = [k for k in ("date", "hours", "address") if k not in col]
    if missing:
        raise ValueError(
            "Не найдены обязательные колонки в выгрузке: "
            + ", ".join(missing)
            + ". Нужно чтобы были дата/часы/адрес (адрес может называться 'Сегмент')."
        )

    out: List[ParsedRow] = []
    for r in rows[header_i + 1 :]:
        d = _to_date(r[col["date"]] if col["date"] < len(r) else None)
        h = _to_hours(r[col["hours"]] if col["hours"] < len(r) else None)

        raw_addr_part = str(r[col["address"]] if col["address"] < len(r) else "").strip()
        raw_city_part = str(r[col["city"]] if "city" in col and col["city"] < len(r) else "").strip()

        addr_part_norm = _to_address(raw_addr_part)
        city_part_norm = _to_address(raw_city_part)

        # raw full: "Город, Сегмент"
        addr_raw_full = raw_addr_part
        if raw_city_part and raw_addr_part and not raw_addr_part.strip().casefold().startswith(
            raw_city_part.strip().casefold()
        ):
            addr_raw_full = f"{raw_city_part}, {raw_addr_part}"
        elif raw_city_part and not raw_addr_part:
            addr_raw_full = raw_city_part

        # normalized full: "город, адрес"
        a_norm = addr_part_norm
        if city_part_norm and addr_part_norm and not addr_part_norm.startswith(city_part_norm):
            a_norm = f"{city_part_norm}, {addr_part_norm}"
        elif city_part_norm and not addr_part_norm:
            a_norm = city_part_norm

        if not d or h is None or not a_norm:
            continue
        if h == Decimal("0.00"):
            continue

        out.append(
            ParsedRow(
                key=RowKey(d=d, addr=a_norm),
                addr_raw=addr_raw_full.strip(),
                hours=h,
                raw=tuple(r),
            )
        )

    return out


def parse_xlsx(path: str) -> List[ParsedRow]:
    """
    Авто-детект:
    - если в первой строке много дат => matrix (заказчик)
    - иначе => rowwise (база)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(values_only=True), None)
    if not header:
        return []

    default_year = datetime.now().year
    date_like = sum(1 for x in header if _header_to_date(x, default_year) is not None)
    if date_like >= 10:
        return parse_xlsx_matrix(path)

    return parse_xlsx_rowwise(path)


# -----------------------------
# compare + report
# -----------------------------
def compare(customer_rows: List[ParsedRow], db_rows: List[ParsedRow]) -> str:
    """
    Сравнение:
    - ключ: (дата, нормализованный адрес)
    - значение: часы (если несколько строк на ключ — суммируем)
    """
    # Variant B: безопасная склейка адресов
    customer_rows, db_rows = _apply_safe_address_merge(customer_rows, db_rows)

    def agg(rows: List[ParsedRow]) -> Dict[RowKey, Decimal]:
        m: Dict[RowKey, Decimal] = {}
        for pr in rows:
            m[pr.key] = (m.get(pr.key, Decimal("0.00")) + pr.hours).quantize(Decimal("0.01"))
        return m

    c = agg(customer_rows)
    b = agg(db_rows)

    c_keys = set(c.keys())
    b_keys = set(b.keys())

    only_c = sorted(c_keys - b_keys, key=lambda k: (k.d, k.addr))
    only_b = sorted(b_keys - c_keys, key=lambda k: (k.d, k.addr))

    diff_hours: List[Tuple[RowKey, Decimal, Decimal]] = []
    for k in sorted(c_keys & b_keys, key=lambda k: (k.d, k.addr)):
        if c[k] != b[k]:
            diff_hours.append((k, c[k], b[k]))

    lines: List[str] = []
    lines.append("РЕЗУЛЬТАТ СВЕРКИ")
    lines.append("")
    lines.append(f"Строк (после очистки): заказчик={len(customer_rows)}, база={len(db_rows)}")
    lines.append(f"Уникальных ключей (дата+адрес): заказчик={len(c)}, база={len(b)}")
    lines.append("")
    lines.append(f"Есть у заказчика, нет в базе: {len(only_c)}")
    lines.append(f"Есть в базе, нет у заказчика: {len(only_b)}")
    lines.append(f"Несовпадение часов при одинаковом ключе: {len(diff_hours)}")
    lines.append("")

    def fmt_key(k: RowKey) -> str:
        return f"{k.d:%d.%m.%Y} | {k.addr}"

    if only_c:
        lines.append("ТОЛЬКО У ЗАКАЗЧИКА (первые 50):")
        for k in only_c[:50]:
            lines.append(f"  - {fmt_key(k)} | часы={c[k]}")
        lines.append("")

    if only_b:
        lines.append("ТОЛЬКО В БАЗЕ (первые 50):")
        for k in only_b[:50]:
            lines.append(f"  - {fmt_key(k)} | часы={b[k]}")
        lines.append("")

    if diff_hours:
        lines.append("РАЗНЫЕ ЧАСЫ (первые 50):")
        for k, hc, hb in diff_hours[:50]:
            lines.append(f"  - {fmt_key(k)} | заказчик={hc} | база={hb}")
        lines.append("")

    return "\n".join(lines)


CRITICAL_DIFF_HOURS = Decimal("1.00")  # порог критичности


def build_report_xlsx(customer_rows: List[ParsedRow], db_rows: List[ParsedRow]) -> bytes:
    """
    Excel-отчёт:
    - Report: менеджерский вид (одна строка = один магазин)
    - Differences: детализация по (дата+адрес)
    """
    # Variant B: безопасная склейка адресов
    customer_rows, db_rows = _apply_safe_address_merge(customer_rows, db_rows)

    # --- агрегаторы ---
    def agg_by_key(rows: List[ParsedRow]) -> Dict[RowKey, Decimal]:
        m: Dict[RowKey, Decimal] = {}
        for pr in rows:
            m[pr.key] = (m.get(pr.key, Decimal("0.00")) + pr.hours).quantize(Decimal("0.01"))
        return m

    def agg_by_addr(rows: List[ParsedRow]) -> Dict[str, Decimal]:
        m: Dict[str, Decimal] = {}
        for pr in rows:
            a = pr.key.addr
            m[a] = (m.get(a, Decimal("0.00")) + pr.hours).quantize(Decimal("0.01"))
        return m

    c_by_key = agg_by_key(customer_rows)
    b_by_key = agg_by_key(db_rows)

    c_by_addr = agg_by_addr(customer_rows)
    b_by_addr = agg_by_addr(db_rows)

    # карты "норм.адрес -> адрес как в файле"
    display_addr: Dict[str, str] = {}
    for pr in customer_rows:
        display_addr.setdefault(pr.key.addr, pr.addr_raw)
    for pr in db_rows:
        display_addr.setdefault(pr.key.addr, pr.addr_raw)

    # -----------------------------
    # менеджерский ключ отображения (для склейки дублей в Report)
    # -----------------------------
    def report_key(display: str) -> str:
        s = (display or "").strip().casefold().replace("ё", "е")
        if not s:
            return ""

        # нормализуем запятые/пробелы
        s = s.replace("—", "-").replace("–", "-")
        s = re.sub(r"\s*,\s*", ",", s)
        s = re.sub(r"\s+", " ", s).strip(" ,")

        parts = [p.strip() for p in s.split(",") if p.strip()]

        # если "тюмень, <другой город>, ..." -> выбрасываем "тюмень"
        if len(parts) >= 2 and parts[0] == "тюмень" and not any(ch.isdigit() for ch in parts[1]):
            parts = parts[1:]

        s = ", ".join(parts)

        # выкидываем всё, что менеджеру не нужно и чаще всего расходится
        # пом/офис/кв/подъезд/этаж/пункты
        s = re.sub(r"\b(помещение|пом|офис|оф|квартира|кв|подъезд|п-?зд|этаж|эт|пункт|п)\b\.?\s*[\w./-]+(?:\s*,\s*[\w./-]+)*", "", s)

        # корпус/к/строение/зд/стр (полностью убираем)
        s = re.sub(r"\b(корпус|корп|к)\b\.?\s*[\w./-]+", "", s)
        s = re.sub(r"\b(строение|стр|здание|зд)\b\.?\s*[\w./-]+", "", s)

        # схлопываем списки домов "39/3, 39/4" -> "39"
        s = re.sub(r"\b(\d+)(?:/\d+)?(?:\s*,\s*\1(?:/\d+)?)+\b", r"\1", s)

        # слеши вокруг
        s = re.sub(r"\s*/\s*", "/", s)

        # чистим мусор
        s = re.sub(r"[.;:]", "", s)
        s = re.sub(r"\s+", " ", s)
        s = re.sub(r"\s*,\s*", ", ", s).strip(" ,")

        return s

    # -----------------------------
    # 1) Лист Report (для менеджера) — СКЛЕИВАЕМ по report_key
    # -----------------------------
    tmp_rows: List[Tuple[str, Decimal, Decimal]] = []
    all_addrs = set(c_by_addr.keys()) | set(b_by_addr.keys())
    for addr_norm in all_addrs:
        disp = display_addr.get(addr_norm, addr_norm)
        tmp_rows.append((disp, c_by_addr.get(addr_norm, Decimal("0.00")), b_by_addr.get(addr_norm, Decimal("0.00"))))

    grouped: Dict[str, Dict[str, Any]] = {}
    for disp, c_sum, b_sum in tmp_rows:
        k = report_key(disp)
        if not k:
            k = disp
        if k not in grouped:
            grouped[k] = {"display": disp, "c": Decimal("0.00"), "b": Decimal("0.00")}
        grouped[k]["c"] = (grouped[k]["c"] + c_sum).quantize(Decimal("0.01"))
        grouped[k]["b"] = (grouped[k]["b"] + b_sum).quantize(Decimal("0.01"))

        # выберем “лучшее” отображение: самое короткое (обычно без лишних хвостов)
        if len(disp) < len(grouped[k]["display"]):
            grouped[k]["display"] = disp

    report_rows: List[Tuple[str, Decimal, Decimal, Decimal]] = []
    for k, v in grouped.items():
        c_sum = v["c"]
        b_sum = v["b"]
        diff = (c_sum - b_sum).quantize(Decimal("0.01"))
        report_rows.append((v["display"], c_sum, b_sum, diff))

    # сортировка: сначала самые большие расхождения
    report_rows.sort(key=lambda x: (abs(x[3]), x[0]), reverse=True)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Report"
    ws.append(["адреса", "заказчик", "база", "расхождения"])

    header_fill = PatternFill("solid", fgColor="E6E6E6")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    for addr, c_sum, b_sum, diff in report_rows:
        ws.append([addr, float(c_sum), float(b_sum), float(diff)])
        r = ws.max_row
        ws.cell(row=r, column=2).fill = green_fill
        ws.cell(row=r, column=3).fill = green_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    # -----------------------------
    # 2) Лист Differences (детальный) — как было
    # -----------------------------
    cust_raw_by_key: Dict[RowKey, str] = {}
    for pr in customer_rows:
        cust_raw_by_key.setdefault(pr.key, pr.addr_raw)

    db_raw_by_key: Dict[RowKey, str] = {}
    for pr in db_rows:
        db_raw_by_key.setdefault(pr.key, pr.addr_raw)

    c_keys = set(c_by_key.keys())
    b_keys = set(b_by_key.keys())

    only_c = sorted(c_keys - b_keys, key=lambda k: (k.addr, k.d))
    only_b = sorted(b_keys - c_keys, key=lambda k: (k.addr, k.d))

    diff_hours: List[Tuple[RowKey, Decimal, Decimal]] = []
    for k in sorted(c_keys & b_keys, key=lambda k: (k.addr, k.d)):
        if c_by_key[k] != b_by_key[k]:
            diff_hours.append((k, c_by_key[k], b_by_key[k]))

    rows_out: List[
        Tuple[str, str, str, date, Optional[Decimal], Optional[Decimal], Optional[Decimal], str, bool]
    ] = []

    for k in only_c:
        cust = c_by_key[k]
        rows_out.append((cust_raw_by_key.get(k, ""), "", k.addr, k.d, cust, None, None, "Только у Заказчика", True))

    for k in only_b:
        dbv = b_by_key[k]
        rows_out.append(("", db_raw_by_key.get(k, ""), k.addr, k.d, None, dbv, None, "Только в Базе", True))

    for k, hc, hb in diff_hours:
        delta = (hc - hb).quantize(Decimal("0.01"))
        critical = abs(delta) >= CRITICAL_DIFF_HOURS
        rows_out.append((cust_raw_by_key.get(k, ""), db_raw_by_key.get(k, ""), k.addr, k.d, hc, hb, delta, "Разные часы", critical))

    rows_out.sort(key=lambda x: (x[2], x[3]))

    ws2 = wb.create_sheet("Differences")
    ws2.append([
        "Адрес (как у Заказчика)",
        "Адрес (как в Базе)",
        "Ключ (норм.)",
        "Дата",
        "Часы (Заказчик)",
        "Часы (База)",
        "Дельта",
        "Статус",
        "Критично",
    ])

    red = PatternFill("solid", fgColor="FFCCCC")
    yellow = PatternFill("solid", fgColor="FFF2CC")

    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for addr_c, addr_b, addr_norm, d, hc, hb, delta, status, critical in rows_out:
        ws2.append([
            addr_c or None,
            addr_b or None,
            addr_norm,
            d,
            float(hc) if hc is not None else None,
            float(hb) if hb is not None else None,
            float(delta) if delta is not None else None,
            status,
            "ДА" if critical else "",
        ])

        row_idx = ws2.max_row
        fill = red if critical else yellow
        for col in range(1, 10):
            ws2.cell(row=row_idx, column=col).fill = fill

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:I{ws2.max_row}"

    widths = [55, 55, 55, 12, 16, 14, 10, 18, 10]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


