from typing import List, Optional, Dict, Any
from pathlib import Path
from urllib.parse import urlparse
import csv
import re

from django.contrib.auth import get_user_model
from django.db import transaction

from openpyxl import load_workbook

from app.models import SalesLead, SalesRoundRobinState


# --- Маппинг названий колонок (если CSV/XLSX “нормальный”) ---
COL_MAP = {
    # company
    "company_name": "company_name",
    "Компания": "company_name",
    "Employer": "company_name",
    "Организация": "company_name",

    # vacancy / title
    "vacancy": "vacancy",
    "Vacancy": "vacancy",
    "Вакансия": "vacancy",
    "Название вакансии": "vacancy",
    "Title": "vacancy",
    "Заголовок": "vacancy",
    "Название": "vacancy",

    # url
    "ad_url": "ad_url",
    "URL": "ad_url",
    "Url": "ad_url",
    "Link": "ad_url",
    "href": "ad_url",
    "Ссылка": "ad_url",
    "Ссылка на вакансию": "ad_url",
    "Ссылка на объявление": "ad_url",

    # city / location
    "city": "city",
    "Город": "city",
    "Region": "city",
    "Регион": "city",
    "Location": "city",
    "Местоположение": "city",
    "Адрес": "city",

    # email / phone
    "email": "email",
    "Email": "email",
    "E-mail": "email",
    "Почта": "email",
    "Контактный email": "email",

    "phone": "phone",
    "Phone": "phone",
    "Телефон": "phone",
    "Контактный телефон": "phone",

    # misc
    "comment": "comment",
    "Описание": "comment",
    "Description": "comment",
    "Текст": "comment",

    "work_types": "work_types",
    "Тип работ": "work_types",

    "staff_count": "staff_count",
    "Кол-во": "staff_count",
    "Количество": "staff_count",
}


def _norm_source(val: str) -> Optional[str]:
    v = (val or "").strip().lower()
    if v in ("hh", "hh.ru", "headhunter"):
        return "hh"
    if v in ("avito", "авито"):
        return "avito"
    return None


def _detect_source_from_url(url: str) -> str:
    u = (url or "").lower()
    if "hh.ru" in u:
        return "hh"
    if "avito.ru" in u:
        return "avito"
    return "avito"


def _split_work_types(val: str) -> List[str]:
    raw = (val or "").strip()
    if not raw:
        return []
    return [x.strip() for x in raw.split(",") if x.strip()]


def _guess_work_types(text: str) -> List[str]:
    t = (text or "").strip().lower()
    out: List[str] = []

    def add(name: str):
        if name not in out:
            out.append(name)

    if "груз" in t:
        add("Грузчик")
    if "убор" in t or "клининг" in t:
        add("Уборщик")
    if "торгов" in t or "ртк" in t or "ртз" in t or "выклад" in t:
        add("Работник торгового зала")
    if "комплект" in t:
        add("Комплектовщик")
    if "фасов" in t:
        add("Фасовщик")
    if "сбор" in t:
        add("Сборщик")

    return out


def _normalize_city_from_avito_url(ad_url: str) -> str:
    """
    Avito: /tyumenskaya_oblast/tyumen/... или /moskva/...
    Иногда 1-й сегмент — область, 2-й — город.
    """
    try:
        parts = urlparse(ad_url).path.strip("/").split("/")
        if not parts:
            return "Не указан"

        first = parts[0].replace("_", " ").replace("-", " ").title()

        if len(parts) >= 2:
            f0 = parts[0].lower()
            if "oblast" in f0 or "kray" in f0 or "respublika" in f0:
                return parts[1].replace("_", " ").replace("-", " ").title()

        return first
    except Exception:
        return "Не указан"


def _is_valid_url(url: str) -> bool:
    try:
        p = urlparse(url)
        return bool(p.scheme and p.netloc)
    except Exception:
        return False


def _get_sales_managers() -> List[Any]:
    User = get_user_model()
    return list(User.objects.filter(is_active=True, is_staff=True).order_by("id"))


def _next_manager() -> Optional[Any]:
    managers = _get_sales_managers()
    if not managers:
        return None

    state, _ = SalesRoundRobinState.objects.get_or_create(pk=1)

    if not state.last_manager_id:
        state.last_manager = managers[0]
        state.save(update_fields=["last_manager"])
        return managers[0]

    ids = [m.id for m in managers]
    if state.last_manager_id not in ids:
        state.last_manager = managers[0]
        state.save(update_fields=["last_manager"])
        return managers[0]

    idx = ids.index(state.last_manager_id)
    nxt = managers[(idx + 1) % len(managers)]
    state.last_manager = nxt
    state.save(update_fields=["last_manager"])
    return nxt


def _normalize_headers(headers: List[str]) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for i, h in enumerate(headers):
        key = (h or "").strip()
        mapped = COL_MAP.get(key)
        if mapped:
            idx[mapped] = i
    return idx


def _get_cell(row, idx: Dict[str, int], key: str) -> str:
    i = idx.get(key)
    if i is None or i >= len(row):
        return ""
    v = row[i]
    return (str(v).strip() if v is not None else "")


@transaction.atomic
def import_sales_leads_xlsx(path: str) -> Dict[str, int]:
    """
    Импорт лидов:
    - .csv (в т.ч. “грязный DOM-CSV” Avito/HH)
    - .xlsx (наш шаблон)
    """
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".csv":
        return _import_csv(path)
    return _import_xlsx_our_template(path)


def _import_csv(path: str) -> Dict[str, int]:
    created = 0
    skipped_dup = 0
    skipped_bad = 0

    # читаем CSV (кодировка)
    raw = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                raw = f.read()
            break
        except Exception:
            continue

    if not raw:
        raise ValueError("CSV пустой или не читается")

    # разделитель
    try:
        dialect = csv.Sniffer().sniff(raw[:3000], delimiters=";,")
        delim = dialect.delimiter
    except Exception:
        delim = ";"

    rows = list(csv.reader(raw.splitlines(), delimiter=delim))
    if len(rows) <= 1:
        return {"created": 0, "skipped_dup": 0, "skipped_bad": 0}

    headers = [((h or "").strip()) for h in rows[0]]
    idx = _normalize_headers(headers)

    # если есть нормальные колонки — используем их
    has_structured = ("ad_url" in idx) and (("vacancy" in idx) or ("company_name" in idx))

    # общий regex под ссылки Avito/HH
    url_re = re.compile(r"https?://(?:www\.)?(?:avito\.ru|hh\.ru)/\S+")

    for row in rows[1:]:
        if not row:
            continue

        vacancy = ""
        company_name = ""
        ad_url = ""
        city = ""
        email = ""
        phone = ""
        comment = ""

        if has_structured:
            ad_url = _get_cell(row, idx, "ad_url")
            vacancy = _get_cell(row, idx, "vacancy")
            company_name = _get_cell(row, idx, "company_name")
            city = _get_cell(row, idx, "city")
            email = _get_cell(row, idx, "email")
            phone = _get_cell(row, idx, "phone")
            comment = _get_cell(row, idx, "comment")

            if not vacancy:
                vacancy = company_name

            source = _detect_source_from_url(ad_url)

            if not comment:
                comment = "Импортировано автоматически из HH.ru" if source == "hh" else "Импортировано автоматически из Avito"

            # если город пустой и это Avito — попробуем из URL
            if (not city) and source == "avito":
                city = _normalize_city_from_avito_url(ad_url)

        else:
            # DOM-CSV (Avito/HH): берём ссылку + самый длинный текст строки как вакансию
            texts = [str(c).strip() for c in row if c and len(str(c).strip()) > 5]

            # 1) ссылка
            for cell in row:
                if not cell:
                    continue
                m = url_re.search(str(cell))
                if m:
                    ad_url = m.group(0)
                    break

            if not ad_url:
                skipped_bad += 1
                continue

            source = _detect_source_from_url(ad_url)
            comment = "Импортировано автоматически из HH.ru" if source == "hh" else "Импортировано автоматически из Avito"

            # 2) вакансия — самый длинный текст
            vacancy = max(texts, key=len, default="Линейный персонал")
            vacancy = re.sub(r"\s+", " ", vacancy).strip()

            # если вдруг схватили “мусор” — подстрахуемся
            if len(vacancy) < 3:
                wt = _guess_work_types(" ".join([str(x) for x in row if x]))
                vacancy = wt[0] if wt else "Линейный персонал"

            # 3) город
            if source == "avito":
                city = _normalize_city_from_avito_url(ad_url)
            else:
                city = "Не указан"

            company_name = vacancy
            email = ""
            phone = ""

        # обязательное
        if not ad_url or not vacancy:
            skipped_bad += 1
            continue

        if not _is_valid_url(ad_url):
            skipped_bad += 1
            continue

        # ограничения модели
        ad_url = ad_url[:200]
        vacancy = vacancy[:255]
        company_name = (company_name or vacancy)[:255]
        city = (city or "Не указан")[:100]
        phone = (phone or "")[:50]

        # дедуп по URL
        if SalesLead.objects.filter(ad_url=ad_url).exists():
            skipped_dup += 1
            continue

        work_types = _guess_work_types(vacancy)
        if not work_types:
            work_types = ["Линейный персонал"]

        manager = _next_manager()

        lead = SalesLead(
            company_name=company_name,
            vacancy=vacancy,
            source=source,
            ad_url=ad_url,
            city=city,
            email=email,
            phone=phone,
            work_types=work_types,
            staff_count=None,
            comment=comment,
            status="new",
            manager=manager,
        )
        lead.full_clean()
        lead.save()
        created += 1

    return {"created": created, "skipped_dup": skipped_dup, "skipped_bad": skipped_bad}


def _import_xlsx_our_template(path: str) -> Dict[str, int]:
    wb = load_workbook(path)
    ws = wb.active

    headers = [(c.value or "").strip() for c in ws[1]]

    required = {"company_name", "source", "ad_url", "city", "email", "work_types", "staff_count", "comment"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"В файле нет колонок: {', '.join(sorted(missing))}")

    idx = {h: headers.index(h) for h in headers}

    created = 0
    skipped_dup = 0
    skipped_bad = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        company_name = (row[idx["company_name"]] or "").strip()
        source_raw = (row[idx["source"]] or "").strip()
        ad_url = (row[idx["ad_url"]] or "").strip()
        city = (row[idx["city"]] or "").strip()
        email = (row[idx["email"]] or "").strip()
        work_types_raw = row[idx["work_types"]] or ""
        staff_count = row[idx["staff_count"]]
        comment = (row[idx["comment"]] or "").strip()

        source = _norm_source(source_raw)
        if not source:
            skipped_bad += 1
            continue

        if not company_name or not ad_url or not city:
            skipped_bad += 1
            continue

        if not _is_valid_url(ad_url):
            skipped_bad += 1
            continue

        ad_url = ad_url[:200]

        if SalesLead.objects.filter(ad_url=ad_url).exists():
            skipped_dup += 1
            continue

        manager = _next_manager()

        work_types = _split_work_types(str(work_types_raw))
        if not work_types:
            work_types = ["Линейный персонал"]

        lead = SalesLead(
            company_name=company_name[:255],
            vacancy=company_name[:255],
            source=source,
            ad_url=ad_url,
            city=city[:100],
            email=email,
            phone="",
            work_types=work_types,
            staff_count=int(staff_count) if staff_count not in (None, "") else None,
            comment=comment,
            status="new",
            manager=manager,
        )
        lead.full_clean()
        lead.save()
        created += 1

    return {"created": created, "skipped_dup": skipped_dup, "skipped_bad": skipped_bad}
